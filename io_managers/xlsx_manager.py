import openpyxl

def store_to_excel(excel_filename: str, data_list: list[dict]):
    """
    Store a list of dictionaries (data entries) into the first table of an excel file.
    
    :params str excel_filename: Path to the Excel file
    :params list[dict] data_list: List of dictionaries to be written to the file
    :return: None
    """
    if not data_list:
        return
    
    # Get the field names from the first data item
    data_fields = list(data_list[0].keys())
    try:
        # Try to open an existing Excel file
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active
        # Get existing headers from the first row
        existing_headers = [cell.value for cell in worksheet[1]]
        # Find new headers that are not in the existing file
        new_headers = [field for field in data_fields if field not in existing_headers]
        # Append new headers to header row
        for i, header in enumerate(new_headers, start=len(existing_headers) + 1):
            worksheet.cell(row=1, column=i, value=header)
        # Combine existing and new headers
        column_headers = existing_headers + new_headers
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # Initialize header row
        worksheet.append(data_fields)
        column_headers = data_fields
    
    # Write data rows
    for data_row in data_list:
        worksheet.append([data_row.get(header, "") for header in column_headers])
    
    workbook.save(excel_filename)

def read_from_excel(excel_filename:str , fields=[]):
    """
    Read from the first table of the Excel file.
    
    :params str excel_filename: Path to the Excel file
    :params list[str] fields: List of fields to read. If empty, all fields are read

    """
    try:
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active
        
        header_row = [cell.value for cell in worksheet[1]]
        
        if len(fields) == 0:
            # Unspecified fields, read all fields
            column_indices = range(len(header_row))
            selected_headers = header_row
        else:
            # Read only the specified fields
            column_indices = [header_row.index(col) for col in fields if col in header_row]
            selected_headers = fields

        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {selected_headers[i]: row[column_indices[i]] for i in range(len(column_indices))}
            data.append(row_data)
            
        if len(data) == 0:
            raise ValueError(f"No data found for any specified column(s): \"{fields}\". Either the file \"{excel_filename}\" is empty or none of the column(s) exist.")
        return data
    
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file \"{excel_filename}\" not found. You are likely to read from a non-existent file.")

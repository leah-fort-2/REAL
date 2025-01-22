import openpyxl

def store_to_excel(excel_filename, data_list):
    if not data_list:
        return
    
    try:
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active
        # Get existing headers
        existing_headers = [cell.value for cell in worksheet[1]]
        # Merge existing headers with new headers
        column_headers = list(set(existing_headers + list(data_list[0].keys())))
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        column_headers = data_list[0].keys()
    
    # If new headers are added, update the header row
    if len(column_headers) > len(existing_headers):
        for col, header in enumerate(column_headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
    
    # Write data rows
    for data_row in data_list:
        worksheet.append([data_row.get(header, "") for header in column_headers])
    
    workbook.save(excel_filename)

def read_from_excel(excel_filename, fields=[]):
    try:
        workbook = openpyxl.load_workbook(excel_filename)
        worksheet = workbook.active
        
        header_row = [cell.value for cell in worksheet[1]]
        
        if len(fields) == 0:
            # Unspecified fields, read all fields
            column_indices = range(len(header_row))
            selected_headers = header_row
        else:
            # Read only the specified fields
            column_indices = [header_row.index(col) for col in fields if col in header_row]
            selected_headers = fields

        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_data = {selected_headers[i]: row[column_indices[i]] for i in range(len(column_indices))}
            data.append(row_data)
            
        if len(data) == 0:
            raise ValueError(f"No data found for any specified column(s): \"{fields}\". Either the file \"{excel_filename}\" is empty or none of the column(s) exist.")
        return data
    
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file \"{excel_filename}\" not found. You are likely to read from a non-existent file.")
